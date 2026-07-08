"""
导入脚本：从 Excel 文件读取书籍数据并导入到 SQLite 数据库
数据库: cool-admin-demo/cool.db
表名: knowledge_book
"""

import re
import sqlite3
from datetime import datetime

import openpyxl

# 路径配置
EXCEL_PATH = r'd:\Users\kaifa\WorkBuddy260605\doc-projects\archive\两性心理学_性教育_资源库条目.xlsx'
DB_PATH = r'd:\Users\kaifa\Trae_cn260425\cool-admin-demo\cool.db'

# 时间戳
NOW = datetime.now().strftime('%Y-%m-%d %H:%M:%S')


def extract_rating(text):
    """从豆瓣评分字段提取数字，如 '豆瓣 9.1' -> 9.1, '豆瓣约 8.0' -> 8.0"""
    if not text:
        return None
    text = str(text).strip()
    match = re.search(r'(\d+\.?\d*)', text)
    if match:
        return float(match.group(1))
    return None


def parse_name(full_name):
    """
    解析书名/原名列
    规则：
    - 如果有 /，按 / 拆分，第一部分为 name，第二部分为 originalName
    - 如果没有 /，全部作为 name，originalName 为空
    """
    if not full_name:
        return '', ''
    full_name = str(full_name).strip()
    if '/' in full_name:
        parts = [p.strip() for p in full_name.split('/', 1)]
        return parts[0], parts[1] if len(parts) > 1 else ''
    return full_name, ''


def main():
    # 1. 读取 Excel
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    print(f"数据表: {ws.title}, 总行数: {ws.max_row}, 总列数: {ws.max_column}")

    # 2. 连接数据库
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    # 开始事务
    conn.execute('BEGIN')

    success_count = 0
    skip_count = 0
    row_num = 1  # 行计数，仅用于日志

    for row_idx in range(1, ws.max_row + 1):
        # 检查第 1 列是否为数字，以此判断是否为数据行
        cell_val = ws.cell(row=row_idx, column=1).value
        if cell_val is None:
            continue
        if not isinstance(cell_val, (int, float)):
            continue

        row_num = row_idx  # 实际 Excel 行号

        # 读取各列（列索引从 1 开始）
        priority = ws.cell(row=row_idx, column=2).value  # Col 2: 优先级
        full_name = ws.cell(row=row_idx, column=3).value  # Col 3: 书名/原名
        author = ws.cell(row=row_idx, column=4).value  # Col 4: 作者
        year = ws.cell(row=row_idx, column=5).value  # Col 5: 年份
        country = ws.cell(row=row_idx, column=6).value  # Col 6: 国家
        synopsis = ws.cell(row=row_idx, column=7).value  # Col 7: 内容简介
        background = ws.cell(row=row_idx, column=8).value  # Col 8: 背景故事
        rating_text = ws.cell(row=row_idx, column=9).value  # Col 9: 豆瓣评分
        tags = ws.cell(row=row_idx, column=10).value  # Col 10: 分类（映射到 tags）
        quality = ws.cell(row=row_idx, column=11).value  # Col 11: 质量

        # 解析书名
        name, original_name = parse_name(full_name)

        # 跳过名称为空的行
        if not name:
            print(f"跳过空名称行第 {row_num} 行")
            skip_count += 1
            continue

        # 提取豆瓣评分
        douban_rating = extract_rating(rating_text)

        # 处理字段默认值
        priority = str(priority).strip() if priority else None
        quality = str(quality).strip() if quality else None
        author = str(author).strip() if author else None
        country = str(country).strip() if country else None
        tags = str(tags).strip() if tags else None
        synopsis = str(synopsis).strip() if synopsis else None
        background = str(background).strip() if background else None

        # 年份转为整数
        if year is not None:
            try:
                year = int(year)
            except (ValueError, TypeError):
                year = None

        # 执行 INSERT
        cursor.execute('''
            INSERT INTO knowledge_book 
                (name, originalName, author, year, country, synopsis, 
                 backgroundStory, doubanRating, priority, quality, tags, cover,
                 createTime, updateTime, tenantId)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            name,
            original_name if original_name else None,
            author,
            year,
            country,
            synopsis,
            background,
            douban_rating,
            priority,
            quality,
            tags,
            '[]',  # cover 默认空数组
            NOW,
            NOW,
            1,  # tenantId
        ))

        print(f"正在导入第 {row_num} 行: {name}")
        success_count += 1

    # 提交事务
    conn.commit()
    conn.close()

    print(f"\n成功导入 {success_count} 条记录")
    if skip_count > 0:
        print(f"跳过 {skip_count} 条空名称行")


if __name__ == '__main__':
    main()
